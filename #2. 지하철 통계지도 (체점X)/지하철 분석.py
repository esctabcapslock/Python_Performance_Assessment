#!/usr/bin/env python
# coding: utf-8

# In[1]:


from openpyxl import load_workbook
import math
import re

coord_seat = load_workbook("data/코레일 좌표데이터.xlsx")['수도권 지하철 수정'];

stn_seat = load_workbook("data/교통카드데이터(수도권도시광역철도).xlsx", data_only=True)['OD자료'];


# In[2]:


class Line(): #가장 기본이 됨
    __l_list=set() # 문자열이 저장된, 노선의 모임
    def __init__(self,l_name): #노선의 정보를 지정. 하나의 역이 중복되어있다면 합쳐주는 역할.
        self.l_name=l_name
        if not (l_name in  Line.__l_list): Line.__l_list.add(l_name)
            
    def max_list(self,li): #그림을 그릴때, 필요
        m=0
        for i in li:
            m=max(m,i.pop)
        return m
            
    coord_line = {"서울교통공사 1호선" :"1호선",
"인천 도시철도 2호선" :"인천2호선",
"서울교통공사 2호선 성수지선" :"2호선",
"일산선" :"3호선",
"신분당선" :"신분당선",
"분당선" :"분당선",
"중앙선" :"경의중앙선",
"경춘선" :"경춘선",
"경강선" :"경강선",
"서울교통공사 2호선 신정지선" :"2호선",
"장항선" :"1호선",
"인천 도시철도 1호선" :"인천1호선",
"소사원시선" :"서해선",
"서울교통공사 6호선" :"6호선",
"인천공항 자기부상철도" :"자기부상",
"망우선" :"경춘선",
"용산선" :"경의중앙선",
"안산선" :"4호선",
"경원선" :"1호선",
"인천국제공항철도" :"공항철도",
"경부고속선" :"1호선",
"의정부경전철" :"의정부경전철",
"용인경전철" :"용인에버라인",
"서울교통공사 5호선 마천지선" :"5호선",
"서울교통공사 4호선" :"4호선",
"서울 경전철 우이신설선" :"우이신설선",
"서울교통공사 5호선" :"5호선",
"수인선" :"수인선",
"김포골드라인" :"김포골드라인",
"과천선" :"4호선",
"서울교통공사 2호선" :"2호선",
"서울교통공사 8호선" :"8호선",
"서울시메트로 9호선" :"9호선",
"서울교통공사 3호선" :"3호선",
"경의선" :"경의중앙선",
"서울교통공사 7호선" :"7호선",
"경인선" :"1호선",
"경부선" :"1호선",
"병점기지선":"1호선"      
}
    od_line={"인천2호선" :"인천2호선",
"9호선2단계" :"9호선",
"일산선" :"3호선",
"우이신설선" :"우이신설선",
"신분당선" :"신분당선",
"경기철도" :"신분당선",
"인천1호선" :"인천1호선",
"1호선" :"1호선",
"분당선" :"분당선",
"중앙선" :"경의중앙선",
"경강선" :"경강선",
"경춘선" :"경춘선",
"2호선" :"2호선",
"장항선" :"1호선",
"용인에버라인" :"용인에버라인",
"4호선" :"4호선",
"안산선" :"4호선",
"7호선" :"7호선",
"경원선" :"1호선",
"9호선" :"9호선",
"의정부경전철" :"의정부경전철",
"공항철도 1호선" :"공항철도",
"수인선" :"수인선",
"과천선" :"4호선",
"6호선" :"6호선",
"3호선" :"3호선",
"5호선" :"5호선",
"경의선" :"경의중앙선",
"경인선" :"1호선",
"경부선" :"1호선",
"8호선" :"8호선"
}
    
class Rail(Line): # 역과 역 사이를 잇는 간선역할
    rail_list=[]
    def get_dis(c1,c2): # 좌표를 상ㅅ
        x1,y1,z1=tuple(map(float,c1.split(',')))
        x2,y2,z2=tuple(map(float,c2.split(',')))
        return ((x1-x2)**2+(y2-y1)**2)**0.5 * 6400 / 180 * math.pi
    
    def __init__(self,st1,st2):
        super().__init__(st1.l_name)
        self.dis = Rail.get_dis(st1.coord, st2.coord)
        self.section = [st1,st2]
        self.coord = [list(map(float,st1.coord.split(',')[0:2])),list(map(float,st2.coord.split(',')[0:2]))]
        self.pop = 0
        Rail.rail_list.append(self)
    
    def get_another_stn(self,st1):
        return self.section[1] if st1 == self.section[0] else self.section[0]
        
    def max_list(self):
        m=0
        for i in Rail.rail_list:
            m=max(m,i.pop)
        return m
    
    
class Station(Line):
    __s_list=[]
    __s_dict={}
    __s_trans_dict={}
    
    
    def __init__(self,l_name,s_name,coord): #오버라이딩
        #print(l_name,s_name)
        if (l_name, s_name) in Station.__s_dict: return
        
        super().__init__(l_name) #부모 객체의 노선 중복 체크 기능을 이용
        self.s_name=s_name
        self.coord = coord
        self.dist=-1 # 다익스트라, 갱신할 부분이다
        self.trans=[] # Station 객체가 들어간 목록
        self.next=[] # Rail 객체가 들어간 전역, 다음역 목록
        self.pre=0 #전역
        self.pop=0
        Station.__s_list.append(self)
        Station.__s_dict[(l_name,s_name)]=self
        
        if s_name not in Station.__s_trans_dict: Station.__s_trans_dict[s_name]=self
        elif s_name not in ['양평', '신촌']: #환승역이다. 이름이 같은 양평,신촌 말고
            trans_stn = Station.__s_trans_dict[s_name] #이름이 같은 다른 역
            
            self.trans.extend([trans_stn]+trans_stn.trans) # 그 역과, 그 역의 환승역을 추가
            for stn in [trans_stn]+trans_stn.trans: # 그 역의 환승역 목록
                stn.trans.append(self)

            
    def get_name(self): #출력을 위한 부분.
        return self.l_name, self.s_name
    
    def all_dist_0 (self=0): #간선 정보를 초기화하여, 다익스트라 알고리즘 실행을 도움
        for stn in Station.__s_list:
            stn.dist=-1 #거리 -1
            stn_pre=0
            
    def print_info(self): #오버라이딩함.
        print("현재 역 정보",self.l_name,self.s_name, self.dist, self.trans)
    
            
    def dijkstra(self,stn): # stn에는 노선과 역정보가 튜플로 들어있다...
        Station.all_dist_0() #초기화
        s_stn = Station.__s_dict[stn] #시발역
        s_stn.dist=0
        print("시작",stn,s_stn)
        visit_list=[s_stn.get_name()]
        front=0
        
        while len(visit_list)-front:
            #print('while 출발역=',s_stn.get_name(), '현재역',front,visit_list[front],len(visit_list),front)
            h_stn = Station.__s_dict[visit_list[front]]
            #h_stn.print_info()
            front+=1
            
            for stn in h_stn.trans: # 환승역들을 확인
                #print('\t환승역',stn.get_name(),stn.dist)
                add_dis = 0 if h_stn.l_name == stn.l_name else 4 # 노선이 다른 경우 거리에 3km 추가
            
                if (stn.dist==-1 or stn.dist > add_dis+h_stn.dist): # 방문한적이 없거나, 더 짧은경우 -> 갱신하기
                    
                    #거리갱신
                    #print(front,'\t\t거리갱신',stn.dist , add_dis,h_stn.dist)
                    stn.dist = add_dis+h_stn.dist
                    stn.pre=h_stn # 전역확인
                    visit_list.append(stn.get_name())
                    
            for rail in h_stn.next: #다음역들에서
                stn = rail.get_another_stn(h_stn)
                #print('\t다음역',stn.s_name)
                
                add_dis = rail.dis # 두 역간 거리 생각
                #print(stn,h_stn)
                if (stn.dist==-1 or stn.dist > add_dis+h_stn.dist): # 방문한적이 없거나, 더 짧은경우 -> 갱신하기
                    
                    #거리갱신
                    #print(front,'\t\t거리갱신',stn.dist , add_dis,h_stn.dist)
                    stn.dist = add_dis+h_stn.dist
                    stn.pre=rail # 전역확인
                    visit_list.append(stn.get_name())
    
    def add_node_pop(self,stn,pop): #역 간 간선의 이용객수를 추가해주는 매소드
        
        while stn.dist!=0:
            #print('add_pop',stn.get_name(), stn.pre, stn.dist)
            if type(stn.pre) == Rail: #전역이 그냥
                stn.pre.pop+=pop # 유동인구추가
                stn=stn.pre.get_another_stn(stn)
            else: #전역이 환승역
                stn=stn.pre
                
    def add_stn_pop(self,st1,st2,pop): #역의 이용객수를 추가해주는 매소드
        Station._Station__s_dict[st1].pop+=pop
        Station._Station__s_dict[st2].pop+=pop
        
    def max_list(self):
        m=0
        for i in Station.__s_list:
            m=max(m,i.pop)
        return m


# In[3]:


# 6호선 응암 처리해야한다...
# 경원선 용산-회기도 처리해야함

# 좌표데이터 입력

pre_stn=stn=''
p_l_name=l_name=''
one_side_stn=[ "서울교통공사 6호선-응암", "서울교통공사 6호선-역촌", "서울교통공사 6호선-불광", "서울교통공사 6호선-독바위", "서울교통공사 6호선-연신내", "서울교통공사 6호선-구산"]
for row in coord_seat: # 역 좌표 데이터를 받아 정리해주는 부분
    pre_stn=stn
    p_l_name=l_name
    
    l_name, s_name, coord = row[0].value, row[1].value, row[2].value
    if not l_name: continue 
    
    l_name_j=Rail.coord_line[l_name] # 정규화된 노선명
    if (l_name_j, s_name) not in Station._Station__s_dict:
        stn=Station(l_name_j,s_name,coord)
    else: 
        stn=Station._Station__s_dict[(l_name_j, s_name)]
    
    
    #print(stn)
    if pre_stn and p_l_name == l_name: #같은 노선일 경우
        rail = Rail(pre_stn,stn) #경로 생성
        stn.next.append(rail) #관계생성
        #pre_stn.pre.append(rail)
        
        if l_name+'-'+s_name not in one_side_stn: #일방 관계생성
            pre_stn.next.append(rail) 
            #stn.pre.append(rail)
    


# In[4]:


pre_stn = stn = ('','')

def clear_stn(line,station): 
    station = re.sub(r'\([^)]*\)', '', station) #역이름의 괄호 제거
    
        
    if line=='경원선' and station in [ "용산", "이촌", "서빙고", "한남", "옥수", "응봉", "왕십리", "청량리"]:
        line = '중앙선'
    line = Line.od_line[line] #노선명 정규화
    if station=="총신대입구": station="이수"
    #if station=='지하인천': station='인천'
    
    return line,station

    
for row in stn_seat: #역 방문 데이터를 통해 정리
    if row[1].value=='승차호선': continue
    pre_stn=stn
    line, stantion = (row[1].value,row[3].value)
    if line =="경의선" and stantion in ["김포공항","계양",'검암']: continue #이상한 데이터
    stn = clear_stn(line, stantion)
    
    if stn != pre_stn: # 달라진 경우 -> 다익스트라 후, 각각의 가중치를 노드에 더한다.
        
        Station.dijkstra(0,stn)
    
    
    n_line,n_stantion,pop  = row[4].value, row[6].value, row[8].value # 도착역 정보
    if n_line =="하차미태그": continue
    
        
    n_stn=clear_stn(n_line,n_stantion) # 출발한다
    
    if n_stn!=stn:
        Station.add_node_pop(0,Station._Station__s_dict[n_stn],pop) # 추가하기
        
    Station.add_stn_pop(0,stn,n_stn,pop)


# In[34]:

# 여기서부터는 그림을 그리기 위한 코드 부분으로, 객체지향을 활용하는 쪽으로 초점이 맞추어져 있다.

import folium;
m = folium.Map([37.6795014,126.746565],tiles='Stamen Toner',zoom_start=10);


# In[35]:


# for rail in Rail.rail_list:
#     print('[',end='')
#     print(rail.section[0].get_name(),rail.section[1].get_name(), rail.coord, rail.pop, sep=',',end=' ')
#     print('],')


# In[36]:


# for stn in Station._Station__s_list:
#     print('[',end='')
#     print(list(map(float,stn.coord.split(',')[0:2])), stn.get_name(), stn.pop, sep=',',end=' ')
#     print('],')


# In[37]:


# max_node=Rail.max_list(0)
# max_stn=Station.max_list(0)
# print(max_node,max_stn)


# In[38]:


from math import sin
from math import cos
from math import pi

def cun_color(x):
    x*=pi*2
    abs_0 = lambda x:x if x>0 else 0
    
    b = abs_0(sin(x)) if x>pi/2 else 1
    g= abs(sin(x)) if abs(x-pi) > pi/2 else 1
    r = abs_0(-sin(x)) if x<3/2*pi else 1
    
    return '#'+"%02X%02X%02X"%(tuple(map(int,(r*255,g*255,b*255))))


# In[39]:


max_node=Rail.max_list(0)
max_stn=Station.max_list(0)


def color(x):
    global max_node
    
    high_p, low_p = max_node, 1
    ind = (x-low_p)/(high_p-low_p)
    return cun_color(ind)
    
def rev(x):
    return [x[0][::-1],x[1][::-1]]

cnt=0
for rail in Rail.rail_list:
    if not rail.pop: continue
   
    folium.PolyLine(
        rev(rail.coord),
        weight=(rail.pop/max_node*40) + 1,
        color=color(rail.pop),
        popup=str(round(rail.pop/365,2))

    ).add_to(m)

for stn in Station._Station__s_list:
    if not stn.pop: continue
    
    folium.CircleMarker(
        list(map(float,stn.coord.split(',')[0:2]))[::-1],
        fill_color="red",
        weight=1,
        color="red",
         popup=str(round(stn.pop/365,2))+str(stn.get_name()), 
        radius= math.sqrt(stn.pop/max_stn)*30,
        tooltop=str(stn.pop)
    ).add_to(m);


# In[40]:


m.save('지하철 경로 전체지도.html'); #파일을 저장한다.


# In[ ]:




